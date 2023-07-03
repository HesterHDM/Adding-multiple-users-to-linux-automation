import openpyxl

# Load the source Excel file
workbook = openpyxl.load_workbook('input.xlsx')
worksheet = workbook.active

# Create a new Excel workbook for the output
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Define the column headers for the output
# output_headers = [
#     '-u',
#     '-g',
#     '600',
#     '-G',
#     'java',
#     '-c',
#     '"oracle user"',
#     '-d',
#     '/home/',
#     '-s',
#     '/bin/bash'
# ]

# Write the headers to the output worksheet
#output_worksheet.append(output_headers)

# Read names from the source worksheet and convert them
row_num = 1  # Start from the second row (assuming the first row contains headers)
current_u = 600
current_sheet = 1

while worksheet.cell(row=row_num, column=1).value:
    name = worksheet.cell(row=row_num, column=1).value

    # Construct the converted name
    converted_name = [
        f'-u {current_u}',  # Unique value starting from 600 in ascending order
        '-g 600',
        '-G java',
        '-c "oracle user"',
        f'-d /home/{name}',
        f'-s /bin/bash {name}'
    ]

    # Write the converted name to the output worksheet
    output_worksheet.append(converted_name)

    row_num += 1
    current_u += 1

    if current_u > 60000:
        # Create a new sheet in the output workbook
        current_sheet += 1
        output_worksheet = output_workbook.create_sheet(title=f'Sheet{current_sheet}')
        #output_worksheet.append(output_headers)
        current_u = 600

# Save the output workbook
output_workbook.save('advoutput.xlsx')

print("Conversion complete. Output saved to 'output.xlsx'.")
